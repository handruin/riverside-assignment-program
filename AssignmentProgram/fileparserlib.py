from openpyxl import load_workbook

from AssignmentProgram.Capacity import Capacity
from AssignmentProgram.Member import Member, Affiliate
from AssignmentProgram.Zipcode import Zipcode


def parse_mass_health_excel():
    # Temporary way to obscure name of file being read in for privacy concerns.
    with open('parse_masshealth_file_name.txt') as f:
        file_name = f.read()
    return load_workbook(file_name).active


def parse_zipcode_excel():
    with open('parse_zipcode_file_name.txt') as f:
        file_name = f.read()
    return load_workbook(file_name).active


def parse_capacity_excel():
    with open('parse_capacity_file_name.txt') as f:
        file_name = f.read()
    return load_workbook(file_name).active


def member_exists(all_members, new_member):
    """
    Look through all parsed members and return True when an identicle member is found based on their last, first name,
    middle initial, date of birth, and medicaid ID.
    :param all_members:
    :param new_member:
    :return: Boolean
    """
    for member in all_members:
        if member.medicaid_id == new_member.medicaid_id \
                and member.last_name == new_member.last_name \
                and member.first_name == new_member.first_name \
                and member.middle_initial == new_member.middle_initial \
                and member.date_of_birth == new_member.date_of_birth:
            return True
    return False


def member_has_different_medicaid_id(all_members, new_member):
    """
    Look through all members and find matching member criteria such as name and date of birth but the medicaid ID is
    different.
    :param all_members:
    :param new_member:
    :return: List
    """
    members_different_id = []
    for member in all_members:
        if member.last_name == new_member.last_name \
                and member.first_name == new_member.first_name \
                and member.middle_initial == new_member.middle_initial \
                and member.date_of_birth == new_member.date_of_birth \
                and member.medicaid_id != new_member.medicaid_id:
            members_different_id.append(member)
    return members_different_id


def get_existing_member(all_members, new_member):
    """
    Look through all members and return the member object if it exists.
    :param all_members:
    :param new_member:
    :return: Member
    """
    for member in all_members:
        if member.medicaid_id == new_member.medicaid_id \
                and member.last_name == new_member.last_name \
                and member.first_name == new_member.first_name \
                and member.middle_initial == new_member.middle_initial \
                and member.date_of_birth == new_member.date_of_birth:
            return member
    return None


def get_record(row, column, active_sheet):
    """
    Helper function for extracting a record from an excel sheet.
    :param row:
    :param column:
    :param active_sheet:
    :return:
    """
    cell_name = "{}{}".format(column, row)
    if active_sheet[cell_name].value:
        if not isinstance(active_sheet[cell_name].value, int):
            val = active_sheet[cell_name].value.lower()
        else:
            val = active_sheet[cell_name].value
        return val


def get_record_and_priority(row, column, active_sheet):
    cell_name = "{}{}".format(column, row)
    priority = False
    val = None
    if active_sheet[cell_name].value:
        if not isinstance(active_sheet[cell_name].value, int):
            val = active_sheet[cell_name].value.lower()
        else:
            val = active_sheet[cell_name].value
        # Checking for the color Yellow used in the zip code excel sheet to indicate priority
        if active_sheet[cell_name].fill.start_color.index == 'FFFFFF00':
            priority = True
    return {"record": [val, priority]}


def is_empty_member(member):
    if member.medicaid_id is None \
            and member.last_name is None \
            and member.first_name is None \
            and member.middle_initial is None \
            and member.date_of_birth is None \
            and member.identification_flag is None:

        for affiliate_obj in member.affiliates:
            if is_affiliate_empty(affiliate_obj):
                return True
    return False


def is_affiliate_empty(affiliate_obj):
    if affiliate_obj.affiliate_name is None \
            and affiliate_obj.is_accs is False \
            and affiliate_obj.is_pcp is False \
            and affiliate_obj.is_cbfs is False \
            and affiliate_obj.is_pact is False \
            and affiliate_obj.is_respite_or_css is False \
            and affiliate_obj.is_out_patient is False \
            and affiliate_obj.is_day_treatment is False \
            and affiliate_obj.is_csp is False \
            and affiliate_obj.is_emergency_svs is False \
            and affiliate_obj.is_ltss is False \
            and affiliate_obj.assigned_to is None:
        return True
    return False


def is_zipcode_empty(zipcode):
    if zipcode.city_town is None \
            and zipcode.zipcode is None \
            and zipcode.ap_lynn is None \
            and zipcode.ap_nsmha is None \
            and zipcode.ap_edinburg is None \
            and zipcode.ap_riverside is None \
            and zipcode.ap_uphams is None \
            and zipcode.ap_dimock is None \
            and zipcode.ap_brookline is None:
        return True
    return False


def is_capacity_empty(capacity):
    if capacity.ap is None \
            and capacity.capacity is None:
        return True
    return False


def update_member_affiliates(all_members, new_member):
    member = get_existing_member(all_members, new_member)
    for affiliate in new_member.affiliates:
        member.add_affiliate(affiliate)


def process_members():
    active_sheet = parse_mass_health_excel()
    obj_members = []

    member_attributes = {
        "medicaid_id": "A",
        "member_last_name": "B",
        "member_first_name": "C",
        "member_middle_initial": "D",
        "member_date_of_birth": "F",
        "residential_address_zipcode_1": "W",
        "identification_flag": "AT",
        "affiliate": "AX",
        "accs": "AY",
        "pcp": "AZ",
        "cbfs": "BA",
        "pact": "BB",
        "respite_or_ccs": "BC",
        "out_patient": "BD",
        "day_treatment": "BE",
        "csp": "BF",
        "emergency_svs": "BG",
        "ltts": "BH",
        "assigned_to": "BI"
    }

    for row in range(2, active_sheet.max_row + 1):
        new_obj_member = Member(get_record(row, member_attributes['medicaid_id'], active_sheet),
                                get_record(row, member_attributes['member_last_name'], active_sheet),
                                get_record(row, member_attributes['member_first_name'], active_sheet),
                                get_record(row, member_attributes['member_middle_initial'], active_sheet),
                                get_record(row, member_attributes['member_date_of_birth'], active_sheet),
                                get_record(row, member_attributes['residential_address_zipcode_1'], active_sheet),
                                get_record(row, member_attributes['identification_flag'], active_sheet))

        new_obj_affiliate = Affiliate(get_record(row, member_attributes['affiliate'], active_sheet))
        new_obj_affiliate.is_accs = bool(int(get_record(row, member_attributes['accs'], active_sheet) or 0))
        new_obj_affiliate.is_pcp = bool(int(get_record(row, member_attributes['pcp'], active_sheet) or 0))

        new_obj_affiliate.is_cbfs = bool(int(get_record(row, member_attributes['cbfs'], active_sheet) or 0))
        new_obj_affiliate.is_pact = bool(int(get_record(row, member_attributes['pact'], active_sheet) or 0))
        new_obj_affiliate.is_respite_or_ccs = bool(
            int(get_record(row, member_attributes['respite_or_ccs'], active_sheet) or 0))
        new_obj_affiliate.is_out_patient = bool(
            int(get_record(row, member_attributes['out_patient'], active_sheet) or 0))
        new_obj_affiliate.is_day_treatment = bool(
            int(get_record(row, member_attributes['day_treatment'], active_sheet) or 0))
        new_obj_affiliate.is_csp = bool(int(get_record(row, member_attributes['csp'], active_sheet) or 0))
        new_obj_affiliate.is_emergency_svs = bool(
            int(get_record(row, member_attributes['emergency_svs'], active_sheet) or 0))
        new_obj_affiliate.is_ltts = bool(int(get_record(row, member_attributes['ltts'], active_sheet) or 0))
        new_obj_affiliate.assigned_to = get_record(row, member_attributes['assigned_to'], active_sheet)

        new_obj_member.add_affiliate(new_obj_affiliate)

        if not is_empty_member(new_obj_member):
            if member_exists(obj_members, new_obj_member):
                update_member_affiliates(obj_members, new_obj_member)
            else:
                obj_members.append(new_obj_member)
    return obj_members


def process_zipcodes():
    active_sheet = parse_zipcode_excel()
    obj_zipcodes = []

    zipcode_base_params = {
        "city_town": "A",
        "zipcode": "B"
    }

    zipcode_attributes = {
        "lynn": "C",
        "nsmha": "D",
        "edinburg": "E",
        "riverside": "F",
        "uphams": "G",
        "dimock": "H",
        "brookline": "I"
    }

    for row in range(2, active_sheet.max_row + 1):
        new_zipcode_obj = Zipcode(get_record(row, zipcode_base_params['city_town'], active_sheet),
                                  get_record(row, zipcode_base_params['zipcode'], active_sheet))

        new_zipcode_obj.ap_lynn = get_record_and_priority(row, zipcode_attributes['lynn'], active_sheet)
        new_zipcode_obj.ap_nsmha = get_record_and_priority(row, zipcode_attributes['nsmha'], active_sheet)
        new_zipcode_obj.ap_edinburg = get_record_and_priority(row, zipcode_attributes['edinburg'], active_sheet)
        new_zipcode_obj.ap_riverside = get_record_and_priority(row, zipcode_attributes['riverside'], active_sheet)
        new_zipcode_obj.ap_uphams = get_record_and_priority(row, zipcode_attributes['uphams'], active_sheet)
        new_zipcode_obj.ap_dimock = get_record_and_priority(row, zipcode_attributes['dimock'], active_sheet)
        new_zipcode_obj.ap_brookline = get_record_and_priority(row, zipcode_attributes['brookline'], active_sheet)
        if not is_zipcode_empty(new_zipcode_obj):
            obj_zipcodes.append(new_zipcode_obj)

    return obj_zipcodes


def process_capacity():
    active_sheet = parse_capacity_excel()
    obj_capacitys = []

    capacity_base_params = {
        "affiliate": "A",
        "capacity": "B"
    }

    for row in range(2, active_sheet.max_row + 1):
        new_capacity_obj = Capacity(get_record(row, capacity_base_params['affiliate'], active_sheet),
                                    get_record(row, capacity_base_params['capacity'], active_sheet))

        if not is_capacity_empty(new_capacity_obj):
            obj_capacitys.append(new_capacity_obj)
    return obj_capacitys


def process_rules(member):
    rule_set_accs(member)


def rule_set_accs(member):
    for affiliate in member.affiliates:
        if affiliate.is_accs is False \
                and affiliate.affiliate_name is None \
                and affiliate.is_cbfs is True:
            if member.identification_flag.lower() == "accs":
                # TODO - change to logger instead of printing to stdout
                print("Rule met and set accs to True for {0}".format(member.first_name))
                affiliate.is_accs = True


all_zipcodes = process_zipcodes()
for zipcode in all_zipcodes:
    # print("city: {0} | zipcode: {1}".format(zipcode.city_town, zipcode.zipcode))
    print("Is zipcode: {0} priority: {1}".format(zipcode.zipcode, zipcode.ap_priority))

all_capacitys = process_capacity()
for capacity in all_capacitys:
    print("ap: {} | capacity: {}".format(capacity.ap, capacity.capacity))

all_members = process_members()
for mem in all_members:
    process_rules(mem)
    for affiliate in mem.affiliates:
        print("{0} - affiliate name: {1}, is_accs: {2}, zipcode: {3}".format(mem.first_name, affiliate.affiliate_name, affiliate.is_accs, mem.residential_address_zipcode_1))

exit(0)

# Rules:
# if accs is None and affiliate is None and cbfs is not None:
# if identification_flag.lower() == "ACCS".lower()
#   set accs = 1
