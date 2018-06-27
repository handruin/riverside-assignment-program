import os

from openpyxl import load_workbook
from openpyxl.styles import Alignment

from AssignmentProgram.Rule import ClientZipcodeInAPServiceAreaRule, RuleProcessing, \
    CurrentACCSclientAPorganizationRule, \
    SetACCSRule, FailedAssignmentRule, CurrentPCPClientAPOrganizationRule, FormerCBFSWithoutTransitionACCSRule, \
    ClientOfBehavioralServiceAPRule, ClientOfLTSSatAPRule, ClientNoPriorHistoryZipcodeCapacityMatchRule
from AssignmentProgram.Capacity import Capacity, CapacityManager
from AssignmentProgram.Member import Member, Affiliate
from AssignmentProgram.Zipcode import Zipcode, ZipcodeManager
import logging
from logging.config import dictConfig

logging_config = dict(
    version=1,
    formatters={
        'f': {'format':
              '%(asctime)s %(name)-12s %(levelname)-8s %(message)s'}
        },
    handlers={
        'h': {'class': 'logging.StreamHandler',
              'formatter': 'f',
              'level': logging.DEBUG}
        },
    root={
        'handlers': ['h'],
        'level': logging.DEBUG,
        },
)

dictConfig(logging_config)
logger = logging.getLogger()
handler = logging.FileHandler('assignment_program.log')
logger.addHandler(handler)


def load_excel_workbook(workbook_location):
    if not os.path.exists(workbook_location):
        raise FileNotFoundError("Error: Unable to locate file: {0}".format(workbook_location))
    return load_workbook(workbook_location).active


def parse_capacity_excel():
    # Temporary way to obscure name of file being read in for privacy concerns.
    # This simple text file should only contain a single line with the full path to the Excel spreadsheet.
    with open('parse_capacity_file_name.txt') as f:
        file_name = f.read()
    return load_workbook(file_name).active


def member_exists(all_members, new_member):
    """
    Look through all parsed members and return True when an identicle member is found based on their last, first name,
    middle initial, date of birth, and medicaid ID.
    :param all_members:
    :param new_member:
    :return: Boolean
    """
    for member in all_members:
        if member.medicaid_id == new_member.medicaid_id \
                and member.last_name == new_member.last_name \
                and member.first_name == new_member.first_name \
                and member.middle_initial == new_member.middle_initial \
                and member.date_of_birth == new_member.date_of_birth:
            return True
    return False


def member_has_different_medicaid_id(all_members, new_member):
    """
    Look through all members and find matching member criteria such as name and date of birth but the medicaid ID is
    different.
    :param all_members:
    :param new_member:
    :return: List
    """
    members_different_id = []
    for member in all_members:
        if member.last_name == new_member.last_name \
                and member.first_name == new_member.first_name \
                and member.middle_initial == new_member.middle_initial \
                and member.date_of_birth == new_member.date_of_birth \
                and member.medicaid_id != new_member.medicaid_id:
            members_different_id.append(member)
    return members_different_id


def get_existing_member(all_members, new_member):
    """
    Look through all members and return the member object if it exists.
    :param all_members:
    :param new_member:
    :return: Member
    """
    for member in all_members:
        if member.medicaid_id == new_member.medicaid_id \
                and member.last_name == new_member.last_name \
                and member.first_name == new_member.first_name \
                and member.middle_initial == new_member.middle_initial \
                and member.date_of_birth == new_member.date_of_birth:
            return member
    return None


def get_record(row, column, active_sheet):
    """
    Helper function for extracting a record from an excel sheet.
    :param row:
    :param column:
    :param active_sheet:
    :return:
    """
    cell_name = "{}{}".format(column, row)
    if active_sheet[cell_name].value:
        if not isinstance(active_sheet[cell_name].value, int):
            val = active_sheet[cell_name].value.lower()
        else:
            val = active_sheet[cell_name].value
        return val


def get_record_cell(row, column, active_sheet):
    """
    Helper function for extracting a record from an excel sheet.
    :param row:
    :param column:
    :param active_sheet:
    :return: Cell
    """
    cell_name = "{}{}".format(column, row)
    if active_sheet[cell_name]:
        return active_sheet[cell_name]


def get_record_and_priority(row, column, active_sheet):
    """
    Helper function used to get a record from an Excel workbook.
    :param row: String
    :param column: String
    :param active_sheet: Workbook
    :return: Dict
    """
    cell_name = "{}{}".format(column, row)
    priority = False
    val = None
    if active_sheet[cell_name].value:
        cell_val = active_sheet[cell_name].value
        # strip out cells with empty strings so they aren't used as a value.
        if cell_val.strip():
            if not isinstance(cell_val, int):
                val = cell_val.lower()
            else:
                val = cell_val
        # Checking for the color Yellow used in the zip code excel sheet to indicate priority
        if active_sheet[cell_name].fill.start_color.index == 'FFFFFF00':
            priority = True
    return {"record": [val, priority]}


def is_empty_member(member):
    """
    Check for an empty Member object to detect for empty Excel rows.
    :param member: Member
    :return: Boolean
    """
    if member.medicaid_id is None \
            and member.last_name is None \
            and member.first_name is None \
            and member.middle_initial is None \
            and member.date_of_birth is None \
            and member.identification_flag is None:

        for affiliate_obj in member.affiliates:
            if is_affiliate_empty(affiliate_obj):
                return True
    return False


def is_affiliate_empty(affiliate_obj):
    """
    Check for an empty Affiliate object to detect for empty Excel rows.
    :param affiliate_obj: Affiliate
    :return: Boolean
    """
    if affiliate_obj.affiliate_name is None \
            and affiliate_obj.is_accs is False \
            and affiliate_obj.is_pcp is False \
            and affiliate_obj.is_cbfs is False \
            and affiliate_obj.is_pact is False \
            and affiliate_obj.is_respite_or_css is False \
            and affiliate_obj.is_out_patient is False \
            and affiliate_obj.is_day_treatment is False \
            and affiliate_obj.is_csp is False \
            and affiliate_obj.is_emergency_svs is False \
            and affiliate_obj.is_ltss is False \
            and affiliate_obj.assigned_to is None:
        return True
    return False


def is_zipcode_empty(zipcode):
    """
    Check for an empty Zipcode object to detect for empty Excel rows.
    :param zipcode: Zipcode
    :return: Boolean
    """
    if zipcode.city_town is None \
            and zipcode.zipcode is None \
            and zipcode.ap_lynn is None \
            and zipcode.ap_nsmha is None \
            and zipcode.ap_edinburg is None \
            and zipcode.ap_riverside is None \
            and zipcode.ap_uphams is None \
            and zipcode.ap_dimock is None \
            and zipcode.ap_brookline is None:
        return True
    return False


def is_capacity_empty(capacity):
    """
    Check for an empty Capacity object to detect for empty Excel rows.
    :param capacity: Capacity
    :return: Boolean
    """
    if capacity.ap is None \
            and capacity.capacity is None:
        return True
    return False


def update_member_affiliates(all_members, new_member):
    """
    Update the member affiliates with a new member data.
    :param all_members: List
    :param new_member: List
    :return: None
    """
    member = get_existing_member(all_members, new_member)
    for affiliate in new_member.affiliates:
        member.add_affiliate(affiliate)


def process_members(member_file_location):
    """
    Process all member data from an Excel workbook into Member objects.
    :param member_file_location: List
    :return: List
    """
    active_sheet = load_excel_workbook(member_file_location)
    obj_members = []

    member_attributes = {
        "medicaid_id": "A",
        "member_last_name": "B",
        "member_first_name": "C",
        "member_middle_initial": "D",
        "member_date_of_birth": "F",
        "residential_address_zipcode_1": "W",
        "identification_flag": "AT",
        "affiliate": "AX",
        "accs": "AY",
        "pcp": "AZ",
        "cbfs": "BA",
        "pact": "BB",
        "respite_or_ccs": "BC",
        "out_patient": "BD",
        "day_treatment": "BE",
        "csp": "BF",
        "emergency_svs": "BG",
        "ltts": "BH",
        "assigned_to": "BI"
    }

    for row in range(2, active_sheet.max_row + 1):
        new_obj_member = Member(get_record(row, member_attributes['medicaid_id'], active_sheet),
                                get_record(row, member_attributes['member_last_name'], active_sheet),
                                get_record(row, member_attributes['member_first_name'], active_sheet),
                                get_record(row, member_attributes['member_middle_initial'], active_sheet),
                                get_record(row, member_attributes['member_date_of_birth'], active_sheet),
                                get_record(row, member_attributes['residential_address_zipcode_1'], active_sheet),
                                get_record(row, member_attributes['identification_flag'], active_sheet))

        new_obj_affiliate = Affiliate(get_record(row, member_attributes['affiliate'], active_sheet))
        new_obj_affiliate.is_accs = bool(int(get_record(row, member_attributes['accs'], active_sheet) or 0))
        new_obj_affiliate.is_pcp = bool(int(get_record(row, member_attributes['pcp'], active_sheet) or 0))

        new_obj_affiliate.is_cbfs = bool(int(get_record(row, member_attributes['cbfs'], active_sheet) or 0))
        new_obj_affiliate.is_pact = bool(int(get_record(row, member_attributes['pact'], active_sheet) or 0))
        new_obj_affiliate.is_respite_or_ccs = bool(
            int(get_record(row, member_attributes['respite_or_ccs'], active_sheet) or 0))
        new_obj_affiliate.is_out_patient = bool(
            int(get_record(row, member_attributes['out_patient'], active_sheet) or 0))
        new_obj_affiliate.is_day_treatment = bool(
            int(get_record(row, member_attributes['day_treatment'], active_sheet) or 0))
        new_obj_affiliate.is_csp = bool(int(get_record(row, member_attributes['csp'], active_sheet) or 0))
        new_obj_affiliate.is_emergency_svs = bool(
            int(get_record(row, member_attributes['emergency_svs'], active_sheet) or 0))
        new_obj_affiliate.is_ltts = bool(int(get_record(row, member_attributes['ltts'], active_sheet) or 0))
        new_obj_affiliate.assigned_to = get_record(row, member_attributes['assigned_to'], active_sheet)

        new_obj_member.add_affiliate(new_obj_affiliate)

        if not is_empty_member(new_obj_member):
            if member_exists(obj_members, new_obj_member):
                update_member_affiliates(obj_members, new_obj_member)
            else:
                obj_members.append(new_obj_member)
    return obj_members


def process_zipcodes(zipcode_file_location):
    """
    Process all Zipcode data from an Excel workbook into Zipcode objects.
    :param zipcode_file_location: List
    :return: List
    """
    active_sheet = load_excel_workbook(zipcode_file_location)
    obj_zipcodes = []

    zipcode_base_params = {
        "city_town": "A",
        "zipcode": "B"
    }

    zipcode_attributes = {
        "lynn": "C",
        "nsmha": "D",
        "edinburg": "E",
        "riverside": "F",
        "uphams": "G",
        "dimock": "H",
        "brookline": "I"
    }

    for row in range(2, active_sheet.max_row + 1):
        new_zipcode_obj = Zipcode(get_record(row, zipcode_base_params['city_town'], active_sheet),
                                  get_record(row, zipcode_base_params['zipcode'], active_sheet))

        new_zipcode_obj.ap_lynn = get_record_and_priority(row, zipcode_attributes['lynn'], active_sheet)
        new_zipcode_obj.ap_nsmha = get_record_and_priority(row, zipcode_attributes['nsmha'], active_sheet)
        new_zipcode_obj.ap_edinburg = get_record_and_priority(row, zipcode_attributes['edinburg'], active_sheet)
        new_zipcode_obj.ap_riverside = get_record_and_priority(row, zipcode_attributes['riverside'], active_sheet)
        new_zipcode_obj.ap_uphams = get_record_and_priority(row, zipcode_attributes['uphams'], active_sheet)
        new_zipcode_obj.ap_dimock = get_record_and_priority(row, zipcode_attributes['dimock'], active_sheet)
        new_zipcode_obj.ap_brookline = get_record_and_priority(row, zipcode_attributes['brookline'], active_sheet)
        if not is_zipcode_empty(new_zipcode_obj):
            obj_zipcodes.append(new_zipcode_obj)

    return obj_zipcodes


def process_capacity(capacity_file_location):
    """
    Process all Capacity data from an Excel workbook into Capacity objects.
    :return: List
    """
    active_sheet = load_excel_workbook(capacity_file_location)
    obj_capacitys = []

    capacity_base_params = {
        "affiliate": "A",
        "capacity": "B"
    }

    for row in range(2, active_sheet.max_row + 1):
        new_capacity_obj = Capacity(get_record(row, capacity_base_params['affiliate'], active_sheet),
                                    get_record(row, capacity_base_params['capacity'], active_sheet))

        if not is_capacity_empty(new_capacity_obj):
            obj_capacitys.append(new_capacity_obj)
    return obj_capacitys


def process_rules(all_members, capacity_manager, zipcode_manager):
    """
    Establish a Rule Processor and assign new Rules to be parsed.
    :param all_members: List
    :param capacity_manager: List
    :param zipcode_manager: List
    :return: None
    """
    for member in all_members:
        rp = RuleProcessing(member, capacity_manager, zipcode_manager)
        rp.assign_rule(ClientZipcodeInAPServiceAreaRule())
        rp.assign_rule(CurrentACCSclientAPorganizationRule())
        rp.assign_rule(SetACCSRule())  # TODO Figure out how this rule fits into the cascade
        rp.assign_rule(CurrentPCPClientAPOrganizationRule())
        rp.assign_rule(FormerCBFSWithoutTransitionACCSRule())
        rp.assign_rule(ClientOfBehavioralServiceAPRule())
        rp.assign_rule(ClientOfLTSSatAPRule())
        rp.assign_rule(ClientNoPriorHistoryZipcodeCapacityMatchRule())
        rp.assign_rule(FailedAssignmentRule())
        rp.process()

        for rule_exception in rp.member_rule_exceptions:
            message = "Medicaid ID: {0} | " \
                      "Member Zipcode: {1} | " \
                      "Message {2} | " \
                      "Rule Class: {3} | " \
                      "Error Type Class: {4}".format(rule_exception.member.medicaid_id,
                                                     rule_exception.member.residential_address_zipcode_1,
                                                     rule_exception.exception_message,
                                                     rule_exception.rule.__class__.__name__,
                                                     rule_exception.assignment_error_type.__class__.__name__)
            logger.debug(message)


def generate_zipcodes_manager(file_location):
    all_zipcodes = process_zipcodes(file_location)
    return ZipcodeManager(all_zipcodes)


def generate_capacities_manager(file_location):
    all_capacities = process_capacity(file_location)
    return CapacityManager(all_capacities)


def generate_members(file_location):
    all_members = process_members(file_location)
    return all_members


# zipcode_manager = generate_zipcodes_manager()
# capacities_manager = generate_capacities_manager()
# all_members = generate_members()
#
# process_rules(all_members, capacities_manager, zipcode_manager)

def generate_stats(all_members):
    count_unassigned = []
    total_count = []
    members_with_multiple_affiliates = []
    for mem in all_members:
        total_count.append(mem)
        if len(mem.affiliates) > 1:
            members_with_multiple_affiliates.append(mem)
        if not mem.is_assigned:
            count_unassigned.append(mem)
            # print("Is {0} assigned: {1}".format(mem.first_name, mem.is_assigned))

    logger.info("Total members: {0}".format(len(total_count)))
    logger.info("Remaining unassigned members: {0}".format(len(count_unassigned)))
    logger.info("Members with multiple affiliate entries: {0}".format(len(members_with_multiple_affiliates)))
    logger.info("="*80)


def update_masshealth_assignments(all_members, masshealth_file_location, mark_duplicates=False):
    duplication_text_marker = "DUPLICATE"
    if not os.path.exists(masshealth_file_location):
        raise FileNotFoundError("Error: Unable to locate file: {0}".format(masshealth_file_location))
    wb = load_workbook(masshealth_file_location)
    active_sheet = wb.active

    member_attributes = {
        "medicaid_id": "A"
    }

    assigned_to_column = {
        "assigned_to": "BI"
    }

    for row in range(2, active_sheet.max_row + 1):
        medicaid_cell = get_record_cell(row, member_attributes['medicaid_id'], active_sheet)
        assigned_to_cell = get_record_cell(row, assigned_to_column['assigned_to'], active_sheet)
        cell_alignment = Alignment(horizontal='center', vertical='center')
        if medicaid_cell.value:
            for member in all_members:
                if member.medicaid_id == medicaid_cell.value:
                    if not member.assignment_written:
                        assigned_to_cell.value = member.get_assigned_affiliate() or ""
                        assigned_to_cell.alignment = cell_alignment
                        member.assignment_written = True
                    else:
                        if mark_duplicates:
                            assigned_to_cell.value = duplication_text_marker
                            assigned_to_cell.alignment = cell_alignment
                    logger.debug("Medicaid: {0} | Assigned: {1}".format(medicaid_cell.value, assigned_to_cell.value))
    wb.save(masshealth_file_location)
